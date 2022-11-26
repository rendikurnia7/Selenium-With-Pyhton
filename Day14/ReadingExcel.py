import openpyxl


#File--->Workbook-->Sheets--->Rows--->Cells

file="C:\\Users\\RezaRendi\\PycharmProjects\\pythonProject\\Selenium with Python\\Day14\\TestFile.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row      #count number of rows in a excel sheet 6
cols=sheet.max_column   #count number of cols in a excel sheet 4


#Reading all Rows and Columns from excel sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='        ')
    print()