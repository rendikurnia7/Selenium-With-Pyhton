import openpyxl

#Write same Data
# file="C:\\Users\\RezaRendi\\PycharmProjects\\pythonProject\\selenium\\Day14\\WriteTest.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active      # (or) sheet=workbook["Data"]   --->get active sheet from excel
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Testing"
#
# workbook.save(file)



#Write multiple data

file="C:\\Users\\RezaRendi\\PycharmProjects\\pythonProject\\selenium\\Day14\\WriteTest1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active      # (or) sheet=workbook["Data"]   --->get active sheet from excel


sheet.cell(1,1).value=123
sheet.cell(1,2).value="Smith"
sheet.cell(1,3).value="Engineer"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="John"
sheet.cell(2,3).value="manager"

sheet.cell(3,1).value=789
sheet.cell(3,2).value="David"
sheet.cell(3,3).value="Developer"

sheet.cell(4,1).value="=SUM(A1:A3)"
workbook.save(file)
